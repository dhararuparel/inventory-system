import io
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from app.models.product import Product
from app.models.inventory import Inventory
from app.models.location import Location

def generate_excel_file(headers, rows, sheet_title, report_name):
    """
    Generates a professionally formatted Excel spreadsheet.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Segoe UI', size=15, bold=True, color='0F172A')
    meta_font = Font(name='Segoe UI', size=9, italic=True, color='64748B')
    header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Segoe UI', size=10, color='1E293B')
    
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo
    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid') # Light slate
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    border_side = Side(border_style='thin', color='E2E8F0')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Write Title
    ws['A1'] = report_name
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 24
    
    # Write Subtitle with Timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A2'] = f"Generated on: {timestamp} | Total Records: {len(rows)}"
    ws['A2'].font = meta_font
    ws.row_dimensions[2].height = 18
    
    # Blank row 3
    ws.row_dimensions[3].height = 8
    
    # Write Table Headers at row 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = cell_border
    ws.row_dimensions[4].height = 24
    
    # Write Data starting at row 5
    for row_idx, row_data in enumerate(rows, 5):
        ws.row_dimensions[row_idx].height = 20
        fill_to_use = zebra_fill if row_idx % 2 == 0 else white_fill
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill_to_use
            cell.border = cell_border
            
            header_name = headers[col_idx - 1].lower()
            
            # Formats & alignments
            if isinstance(val, (int, float)):
                if 'price' in header_name or 'value' in header_name:
                    cell.number_format = '₹#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif 'quantity' in header_name or 'stock' in header_name or 'shortage' in header_name:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:  # skip header metadata
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def parse_and_validate_products(filepath):
    """
    Parses the Product Import excel spreadsheet and runs validation on each row.
    Returns: (is_valid, rows_list, summary)
    """
    try:
        df = pd.read_excel(filepath)
    except Exception as e:
        return False, [], {"error": f"Failed to parse Excel file: {str(e)}"}
    
    # Normalize headers
    normalized_cols = {c: str(c).strip().replace('_', ' ').title() for c in df.columns}
    df = df.rename(columns=normalized_cols)
    
    required_cols = ['Product Code', 'Product Name', 'Category', 'Purchase Price', 'Selling Price']
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        return False, [], {"error": f"Missing required columns in Excel: {', '.join(missing_cols)}"}
        
    rows = []
    seen_codes = set()
    invalid_count = 0
    valid_count = 0
    
    # Load existing codes (case-insensitive checks)
    existing_products = Product.query.with_entities(Product.product_code).all()
    db_codes = {p.product_code.upper() for p in existing_products}
    
    # Optional columns check
    has_brand = 'Brand' in df.columns
    has_size = 'Size' in df.columns
    has_min_stock = 'Minimum Stock' in df.columns
    
    for idx, row in df.iterrows():
        errors = []
        row_num = idx + 2  # Excel is 1-indexed and has header row
        
        # Product Code validation
        code_raw = row.get('Product Code')
        code = str(code_raw).strip().upper() if pd.notna(code_raw) else ''
        
        if not code:
            errors.append("Product Code is required.")
        else:
            # Duplicate check inside Excel
            if code in seen_codes:
                errors.append(f"Duplicate Product Code '{code}' found within the uploaded sheet.")
            else:
                seen_codes.add(code)
                
            # Duplicate check against DB
            if code in db_codes:
                errors.append(f"Product Code '{code}' already exists in database.")
                
        # Product Name validation
        name_raw = row.get('Product Name')
        name = str(name_raw).strip() if pd.notna(name_raw) else ''
        if not name:
            errors.append("Product Name is required.")
            
        # Category validation
        cat_raw = row.get('Category')
        cat = str(cat_raw).strip() if pd.notna(cat_raw) else ''
        if not cat:
            errors.append("Category is required.")
            
        # Purchase Price validation
        p_price_raw = row.get('Purchase Price')
        p_price = None
        if pd.isna(p_price_raw):
            errors.append("Purchase Price is required.")
        else:
            try:
                p_price = float(p_price_raw)
                if p_price < 0:
                    errors.append("Purchase Price cannot be negative.")
            except ValueError:
                errors.append("Purchase Price must be a valid number.")
                
        # Selling Price validation
        s_price_raw = row.get('Selling Price')
        s_price = None
        if pd.isna(s_price_raw):
            errors.append("Selling Price is required.")
        else:
            try:
                s_price = float(s_price_raw)
                if s_price < 0:
                    errors.append("Selling Price cannot be negative.")
                if p_price is not None and s_price < p_price:
                    # Warning rather than error, or soft check
                    pass
            except ValueError:
                errors.append("Selling Price must be a valid number.")
                
        # Minimum Stock validation
        min_stock = 10
        if has_min_stock:
            min_stock_raw = row.get('Minimum Stock')
            if pd.notna(min_stock_raw):
                try:
                    min_stock = int(min_stock_raw)
                    if min_stock < 0:
                        errors.append("Minimum Stock cannot be negative.")
                except ValueError:
                    errors.append("Minimum Stock must be a valid integer.")
                    
        # Brand & Size fallbacks
        brand = str(row.get('Brand')).strip() if has_brand and pd.notna(row.get('Brand')) else ''
        size = str(row.get('Size')).strip() if has_size and pd.notna(row.get('Size')) else ''
        
        row_status = "invalid" if errors else "valid"
        if errors:
            invalid_count += 1
        else:
            valid_count += 1
            
        rows.append({
            "row_number": row_num,
            "product_code": code,
            "name": name,
            "brand": brand,
            "category": cat,
            "size": size,
            "purchase_price": p_price,
            "selling_price": s_price,
            "minimum_stock": min_stock,
            "status": row_status,
            "errors": errors
        })
        
    summary = {
        "total_rows": len(rows),
        "valid_rows": valid_count,
        "invalid_rows": invalid_count,
        "has_errors": invalid_count > 0
    }
    
    return True, rows, summary

def parse_and_validate_inventory(filepath):
    """
    Parses the Inventory Update excel spreadsheet and runs validation.
    Returns: (is_valid, rows_list, summary)
    """
    try:
        df = pd.read_excel(filepath)
    except Exception as e:
        return False, [], {"error": f"Failed to parse Excel file: {str(e)}"}
        
    # Normalize headers
    normalized_cols = {c: str(c).strip().replace('_', ' ').title() for c in df.columns}
    df = df.rename(columns=normalized_cols)
    
    required_cols = ['Product Code', 'Quantity']
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        return False, [], {"error": f"Missing required columns in Excel: {', '.join(missing_cols)}"}
        
    rows = []
    invalid_count = 0
    valid_count = 0
    seen_codes = set()
    
    # Fetch default location "Main Godown"
    godown = Location.query.filter_by(name="Main Godown").first()
    godown_id = godown.id if godown else 1
    
    for idx, row in df.iterrows():
        errors = []
        row_num = idx + 2
        
        # Product Code check
        code_raw = row.get('Product Code')
        code = str(code_raw).strip().upper() if pd.notna(code_raw) else ''
        
        product = None
        current_quantity = 0
        
        if not code:
            errors.append("Product Code is required.")
        else:
            if code in seen_codes:
                errors.append(f"Duplicate Product Code '{code}' in the sheet.")
            else:
                seen_codes.add(code)
                
            # Verify product exists in database
            product = Product.query.filter_by(product_code=code, is_active=True).first()
            if not product:
                errors.append(f"Active Product with Code '{code}' does not exist in the database.")
            else:
                # Fetch current stock level at Main Godown
                inv_record = Inventory.query.filter_by(product_id=product.id, location_id=godown_id).first()
                current_quantity = inv_record.quantity if inv_record else 0
                
        # Quantity validation
        qty_raw = row.get('Quantity')
        new_quantity = None
        if pd.isna(qty_raw):
            errors.append("Quantity is required.")
        else:
            try:
                new_quantity = int(qty_raw)
                if new_quantity < 0:
                    errors.append("Quantity cannot be negative.")
            except ValueError:
                errors.append("Quantity must be a valid integer.")
                
        difference = 0
        if product and new_quantity is not None and not errors:
            difference = new_quantity - current_quantity
            
        row_status = "invalid" if errors else "valid"
        if errors:
            invalid_count += 1
        else:
            valid_count += 1
            
        rows.append({
            "row_number": row_num,
            "product_code": code,
            "product_name": product.name if product else "N/A",
            "current_quantity": current_quantity,
            "new_quantity": new_quantity if new_quantity is not None else 0,
            "difference": difference,
            "status": row_status,
            "errors": errors
        })
        
    summary = {
        "total_rows": len(rows),
        "valid_rows": valid_count,
        "invalid_rows": invalid_count,
        "has_errors": invalid_count > 0
    }
    
    return True, rows, summary
